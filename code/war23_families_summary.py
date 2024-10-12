"""Find groups of family members."""
import os
import numpy as np
import pandas as pd
local = '/home/innereye/alarms/'
if os.path.isdir(local):
    os.chdir(local)
    local = True


db = pd.read_csv('data/oct7database.csv')
## N victims per group
df = pd.read_csv('data/victims_relationship.csv')
group = np.unique(df['group'])
dfn = pd.DataFrame(group, columns=['group'])
for ii in range(len(dfn)):
    pids = df['pid'][df['group'] == group[ii]].values
    name = []
    status = []
    for pid in pids:
        row = np.where(db['pid'] == pid)[0][0]
        name.append(db['שם משפחה'][row])
        status.append(db['Status'][row])
    name = np.sort(np.unique(name))
    name = '; '.join(name)
    killed = sum([1 for k in range(len(status)) if 'killed' in status[k]])
    kidnapped = sum([1 for k in range(len(status)) if 'kidnapped' in status[k]])
    dfn.at[ii, 'name'] = name
    dfn.at[ii, 'victims'] = np.sum(df['group'] == group[ii])
    dfn.at[ii, 'killed'] = killed
    dfn.at[ii, 'kidnapped'] = kidnapped
dfn = dfn.sort_values('victims', ascending=False, ignore_index=True)
dfn.to_excel('~/Documents/group_size.xlsx', index=False)
## parents
together = pd.DataFrame(columns=['relation', 'victims','killed', 'kidnapped'])
columns = ['partners', 'siblings', 'parents to', 'children of', 'grdparents', 'grdchildren', 'other']
for icol, col in enumerate(columns):
    subgroup = df[~df[col].isnull()].copy()
    subgroup = subgroup.reset_index(drop=True)
    victims = len(subgroup)
    killed = 0
    kidnapped = 0
    for ii in range(len(subgroup)):
        pids = [subgroup['pid'][ii]] + [int(float(x)) for x in str(subgroup[col][ii]).split(';')]
        status = []
        for pid in pids:
            row = np.where(db['pid'] == pid)[0][0]
            status.append(db['Status'][row])
        if 'killed' in status[0]:
            if any([True for k in status[1:] if 'killed' in k]):
                killed += 1
        if 'kidnapped' in status[0]:
            if any([True for k in status[1:] if 'kidnapped' in k]):
                kidnapped += 1
    together.at[icol, 'relation'] = col
    together.at[icol, 'victims'] = victims
    together.at[icol, 'killed'] = killed
    together.at[icol, 'kidnapped'] = kidnapped
together.to_excel('~/Documents/victims_together.xlsx', index=False)

##  check victims after oct 7
df = pd.read_csv('data/victims_relationship.csv')
later = df[df['event date'].values > '2023-10-07']
later = later.reset_index(drop=True)
for ii in range(len(later)):
    pids = []
    for col in columns:
        pid = str(later[col][ii])
        if pid != 'nan':
            pids.extend([int(float(x)) for x in pid.split(';')])
            if pids[-1] not in later['pid'].values:
                row = np.where(df['pid'] == pids[-1])[0][0]
                later.loc[len(later)] = df.loc[row]
                print(f"{later['name'][ii]} & {df['name'][row]}")
later = later.sort_values(['group','event date','name'], ignore_index=True)
later.to_excel('~/Documents/after_oct7.xlsx', index=False)
    
        
        
    pids = [int(float(x)) for x in str(subgroup[col][ii]).split(';')]
##

import openpyxl
from openpyxl import load_workbook
from openpyxl.comments import Comment

# excel_file = 'sample.xlsx' 
# wb = load_workbook(excel_file, data_only = True)
# sh = wb['Sheet1']
# # iterate through excel and display data
# for i in range(1, sh.max_row+1):
#     for j in range(1, sh.max_column+1):
#         ## check if the background is yellow
#         if sh.cell(row=i, column=j).fill.start_color.index == 'FFFFFF00':
#             sh.cell(row=i, column=j).comment = Comment("This is a balance account", "author name")
# wb.save('commented_sample.xlsx')


# dfn.at[ii, 'name'] = name
# dfn.at[ii, 'victims'] = np.sum(df['group'] == group[ii])
# dfn.at[ii, 'killed'] = killed
# dfn.at[ii, 'kidnapped'] = kidnapped
