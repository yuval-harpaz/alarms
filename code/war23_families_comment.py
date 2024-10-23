"""Find groups of family members."""
import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.comments import Comment
local = '/home/innereye/alarms/'
if os.path.isdir(local):
    os.chdir(local)
    local = True


# db = pd.read_csv('data/oct7database.csv')
## N victims per group
csv = 'data/victims_relationship.csv'
excel = '/home/innereye/Documents/victims_relationship.xlsx'
df = pd.read_csv(csv)
dfn = df.copy()
df.to_excel(excel, index=False)
wb = load_workbook(excel, data_only = True)
sh = wb['Sheet1']
for icol in range(7,14):
    for row in range(len(df)):
        item = str(df.iloc[row, icol])
        if item != 'nan':
            ids = [int(float(id)) for id in item.split(';')]
            comment = ''
            for id in ids:
                id_row = np.where(df['pid'] == id)[0][0]
                comment += df['name'][id_row]+'; '
            comment = comment[:-2]
            sh.cell(row=row+2, column=icol+1).comment = Comment(comment, author='script')
            dfn.at[row, dfn.columns[icol]] = comment
wb.save('/home/innereye/Documents/oct7kin.xlsx')
dfn.to_csv('~/Documents/oct7kin.csv', index=False)


##
# group = np.unique(df['group'])
# dfn = pd.DataFrame(group, columns=['group'])
# for ii in range(len(dfn)):
#     pids = df['pid'][df['group'] == group[ii]].values
#     name = []
#     status = []
#     for pid in pids:
#         row = np.where(db['pid'] == pid)[0][0]
#         name.append(db['שם משפחה'][row])
#         status.append(db['Status'][row])
#     name = np.sort(np.unique(name))
#     name = '; '.join(name)
#     killed = sum([1 for k in range(len(status)) if 'killed' in status[k]])
#     kidnapped = sum([1 for k in range(len(status)) if 'kidnapped' in status[k]])
#     dfn.at[ii, 'name'] = name
#     dfn.at[ii, 'victims'] = np.sum(df['group'] == group[ii])
#     dfn.at[ii, 'killed'] = killed
#     dfn.at[ii, 'kidnapped'] = kidnapped
# dfn = dfn.sort_values('victims', ascending=False, ignore_index=True)
# dfn.to_excel('~/Documents/group_size.xlsx', index=False)
# ## parents
# together = pd.DataFrame(columns=['relation', 'victims','killed', 'kidnapped'])
# columns = ['partners', 'siblings', 'parents to', 'children of', 'grdparents', 'grdchildren', 'other']
# for icol, col in enumerate(columns):
#     subgroup = df[~df[col].isnull()].copy()
#     subgroup = subgroup.reset_index(drop=True)
#     victims = len(subgroup)
#     killed = 0
#     kidnapped = 0
#     for ii in range(len(subgroup)):
#         pids = [subgroup['pid'][ii]] + [int(float(x)) for x in str(subgroup[col][ii]).split(';')]
#         status = []
#         for pid in pids:
#             row = np.where(db['pid'] == pid)[0][0]
#             status.append(db['Status'][row])
#         if 'killed' in status[0]:
#             if any([True for k in status[1:] if 'killed' in k]):
#                 killed += 1
#         if 'kidnapped' in status[0]:
#             if any([True for k in status[1:] if 'kidnapped' in k]):
#                 kidnapped += 1
#     together.at[icol, 'relation'] = col
#     together.at[icol, 'victims'] = victims
#     together.at[icol, 'killed'] = killed
#     together.at[icol, 'kidnapped'] = kidnapped
# together.to_excel('~/Documents/victims_together.xlsx', index=False)

# ##  check victims after oct 7
# df = pd.read_csv('data/victims_relationship.csv')
# later = df[df['event date'].values > '2023-10-07']
# later = later.reset_index(drop=True)
# for ii in range(len(later)):
#     pids = []
#     for col in columns:
#         pid = str(later[col][ii])
#         if pid != 'nan':
#             pids.extend([int(float(x)) for x in pid.split(';')])
#             if pids[-1] not in later['pid'].values:
#                 row = np.where(df['pid'] == pids[-1])[0][0]
#                 later.loc[len(later)] = df.loc[row]
#                 print(f"{later['name'][ii]} & {df['name'][row]}")
# later = later.sort_values(['group','event date','name'], ignore_index=True)
# later.to_excel('~/Documents/after_oct7.xlsx', index=False)

